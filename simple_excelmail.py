import smtplib

from email.mime.text import MIMEText

from openpyxl import load_workbook

gmail_user = "yourUserName@gmail.com"
gmail_appPassword = "yourAppPassword"

sent_from = ['yourUserName@gmail.com']
to = ['probablyYourUserNameForNow@gmail.com']

def get_invoice():    
    
    excelfile = 'simple_excelmail.xlsx'
    wb = load_workbook(excelfile)
    ws = wb[wb.sheetnames[0]]
    
    invoice = []
    
    for row in ws.iter_rows(row_offset=1):
        for cell in row:
            print(cell.value)
            invoice.append(str(cell.value))
            
    return invoice

data = get_invoice()

text = "{0} owe me {1} dollars, bro".format(data[0], data[3]) 

print(text)

msg = MIMEText(text)

#server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
#server.login(gmail_user, gmail_appPassword)
#server.sendmail(sent_from, to, msg.as_string())
#server.quit()

#tested-works#
