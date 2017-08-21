import smtplib

from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

gmail_user = "yourUserName@gmail.com"
gmail_appPassword = "yourAppPassword"

msg = MIMEMultipart('alternative')
sent_from = ['yourUserName@gmail.com']
msg['Subject'] = 'Taxes: Testing Invoices'
    
#excel-file=data
def list_invoices():    
    
    excelfile = 'example_excel.xlsx'
    wb = load_workbook(excelfile)
    ws = wb[wb.sheetnames[0]]
    
    excel_list = []
    
    for row in ws.iter_rows(row_offset=1):
        for cell in row:
#            print(cell.value)
            excel_list.append(str(cell.value))
            
    return excel_list

# Invoice is Constructor. Builds template rows for html
class Invoice:    
    
    def __init__(self, name, invoice_num, date, amount, notes):
        self.name = name
        self.invoice_num = invoice_num.split(",")
        self.date = date.split(",")
        self.amount = amount.split(",")
        self.notes = notes.split(",")
        self.total = 0
        
        self.template_array = []
#        print(self.notes[0], self.invoice_num, self.to)
        
#        sum amounts for total
        for i in range(0, len(self.invoice_num)):
            self.total += int(self.amount[i])
            
            template = """
                <tr class="left">
                    <td style="padding: 10px; text-align: left;">"""+ self.invoice_num[i] +"""</td>
                    <td style="padding: 10px;">""" + self.date[i] + """</td>
                    <td style="text-align: right; padding: 10px;">"""+ self.amount[i] +"""</td>
                    <td style="padding-left: 20px;">"""+ self.notes[i] +"""</td>
                    </tr>
                """        
            self.template_array.append(template)
        self.total = str(self.total)

        
        
def send_email(to, msg, new_template):
        
    data.reverse()    
    
    for i in range(0, 6):
        data.pop()
           
    try:
        data.reverse()
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(gmail_user, gmail_pword)
        server.sendmail(sent_from, to, msg.as_string())
        
        print("Email Sent To: ", new_template.name)
        print("@: ", to)            
        print("Invoice Numbers: ", new_template.invoice_num)
        print("TOTAL: ", new_template.total, "---------------------------")            
        server.quit()
        
        if(data[0] == 'None'):
            print('END OF LIST')
        else:            
            build_email(data)
    except Exception as e:
        print(e)
        print('Email Failed to Send to: ', new_template.name)
        print("@: ", to)
        print("Invoice Numbers: ", new_template.invoice_num)

data = list_invoices()
#print(data)

def build_email(data):
    
#    print(data[5])
    new_template = Invoice(data[0], data[1], data[2], data[3], data[4])
    
    to = data[5].split(",") #This must be a list i.e. ['someaddress@gmail.com']
    
    text = "Hello, {0}, \n Please find the following invoices attached {1} on these dates {2} for the following amounts {3}. You are encouraged to make payment arrangements now.\n Thank You.\n My Professional Co.".format(data[0], data[1], data[2] data[3])
    
        #use IN-LINE Styling    
    html = """\
    <!DOCTYPE html>
    <html>
        <body>
            <p style="text-align: center"> Hello, """+ new_template.name +""" Hope this email finds you well.</p> 
              
            <p style="text-align: center">Here are your outstanding invoices</p>
            <hr style="width: 500px;">
            <table style="margin-left: auto; margin-right: auto">
                <tr>
                    <th>INVOICE TOTAL:</th>
                    <th style="padding-left: 100px">$"""+ new_template.total +"""</th>
                </tr>
            </table>
            <hr class="width">
            
            <table style="margin-left: auto; margin-right: auto">
                <tr class="left padded">
                    <th style="text-align: left;"> Invoice </th>
                    <th> Date </th>
                    <th style="text-align: right;"> Amount </th>
                    <th style="padding-left: 20px;"> Notes </th>
                </tr>
                """ + ''.join(new_template.template_array) + """
            </table>
            
            <hr style="width: 500px;">
                <table style="margin-left: auto; margin-right: auto; padding: 10px;">
                    <tr>
                        <th> Thank You! </th>
                    </tr>
                    <tr>
                        <th> Some Professional Company LLC </th>
                    </tr>
                </table>
            <hr style="width: 500px;">
            
        </body>
    </html>
    """
    
    part1 = MIMEText(text, 'plain')
    part2 = MIMEText(html, 'html')
    
    
    msg.attach(part1)
    msg.attach(part2)
      
    send_email(to, msg, new_template)  

                               
build_email(data)

